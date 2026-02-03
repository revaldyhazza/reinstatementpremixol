import streamlit as st
import pandas as pd
import numpy as np
from fitter import Fitter
import scipy.stats as stats
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import uuid
import plotly.express as px
import plotly.graph_objects as go

# ============================================================================
# KONFIGURASI HALAMAN
# ============================================================================
st.set_page_config(
    page_title="XoL Reinstatement Pricing", 
    layout="wide", 
    page_icon="📊",
    initial_sidebar_state="expanded"
)

# ============================================================================
# CUSTOM CSS
# ============================================================================
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .info-box {
        background-color: #e7f3ff;
        padding: 1rem;
        color: #0f172a;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #d4edda;
        padding: 1rem;
        color: #0f5132; 
        border-radius: 0.5rem;
        border-left: 4px solid #28a745;
        margin-bottom: 1rem;
    }
    .warning-box {
        background-color: #fff3cd;
        color: #664d03; 
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #ffc107;
        margin-bottom: 1rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
    }
    .stTabs [data-baseweb="tab"] {
        height: 3rem;
        padding: 0 2rem;
    }
    </style>
    """, unsafe_allow_html=True)

# ============================================================================
# FUNGSI UTILITAS
# ============================================================================

def format_currency(value):
    """Format angka menjadi format mata uang Indonesia"""
    if pd.isna(value) or value == '':
        return '-'
    try:
        return f"Rp {int(value):,}".replace(',', '.')
    except:
        return str(value)

def show_info_box(message, box_type="info"):
    """Tampilkan info box dengan styling"""
    if box_type == "info":
        st.markdown(f'<div class="info-box">ℹ️ {message}</div>', unsafe_allow_html=True)
    elif box_type == "success":
        st.markdown(f'<div class="success-box">✅ {message}</div>', unsafe_allow_html=True)
    elif box_type == "warning":
        st.markdown(f'<div class="warning-box">⚠️ {message}</div>', unsafe_allow_html=True)

# ============================================================================
# FUNGSI CORE (sama seperti sebelumnya tapi dengan peningkatan)
# ============================================================================

@st.cache_data
def sesuaikan_distribusi_severitas(data, _cache_key=None):
    """Fitting distribusi severitas dengan progress indicator"""
    distribusi_severitas = ['lognorm', 'gamma', 'weibull_min', 'expon', 
                            'beta', 'pareto', 'invgauss', 'fisk', 
                            'loggamma', 'genpareto', 'erlang', 'cauchy']
    try:
        f = Fitter(data, distributions=distribusi_severitas, timeout=300)
        f.fit()
    except Exception as e:
        raise ValueError(f"Gagal menyesuaikan distribusi severitas: {str(e)}")
    
    parameter_disesuaikan = f.fitted_param
    metrik = []
    
    hist, bin_edges = np.histogram(data, bins='auto', density=True)
    bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
    
    for nama_dist in parameter_disesuaikan:
        try:
            dist = getattr(stats, nama_dist)
            params = parameter_disesuaikan[nama_dist]
            distribusi_disesuaikan = dist(*params)
            
            log_likelihood = np.sum(distribusi_disesuaikan.logpdf(data))
            if not np.isfinite(log_likelihood):
                continue
            
            k = len(params)
            aic = 2 * k - 2 * log_likelihood
            bic = k * np.log(len(data)) - 2 * log_likelihood
            ks_stat, ks_pval = stats.ks_2samp(data, distribusi_disesuaikan.rvs(len(data), random_state=42))
            
            pdf_disesuaikan = distribusi_disesuaikan.pdf(bin_centers)
            pdf_empiris = hist
            panjang_min = min(len(pdf_disesuaikan), len(pdf_empiris))
            pdf_disesuaikan = pdf_disesuaikan[:panjang_min]
            pdf_empiris = pdf_empiris[:panjang_min]
            
            metrik.append({
                'Distribusi': nama_dist,
                'Statistik KS': round(ks_stat, 4),
                'P-Value KS': round(ks_pval, 4),
                'AIC': round(aic, 4),
                'BIC': round(bic, 4),
                'Parameter': params
            })
        except Exception:
            continue
    
    if not metrik:
        raise ValueError("Tidak ada distribusi severitas yang valid.")
    
    df_metrik = pd.DataFrame(metrik)
    return df_metrik.sort_values('AIC').head(10), parameter_disesuaikan

@st.cache_data
def jalankan_simulasi_monte_carlo(jumlah_iterasi, nama_dist_frekuensi, param_frekuensi, 
                                  nama_dist_severitas, param_severitas, _cache_key=None):
    """Simulasi Monte Carlo dengan validasi lebih baik"""
    data_tabel = []
    np.random.seed(42)
    
    # Setup distribusi frekuensi
    if nama_dist_frekuensi == 'poisson':
        dist_frekuensi = stats.poisson
        if param_frekuensi['mu'] <= 0:
            raise ValueError("Parameter mu untuk distribusi Poisson harus positif.")
    elif nama_dist_frekuensi == 'nbinom':
        dist_frekuensi = stats.nbinom
        if param_frekuensi['n'] <= 0 or param_frekuensi['p'] <= 0 or param_frekuensi['p'] >= 1:
            raise ValueError("Parameter n dan p untuk distribusi Negative Binomial tidak valid.")
    elif nama_dist_frekuensi == 'geom':
        dist_frekuensi = stats.geom
        if param_frekuensi['p'] <= 0 or param_frekuensi['p'] >= 1:
            raise ValueError("Parameter p untuk distribusi Geometric tidak valid.")
    else:
        raise ValueError("Distribusi frekuensi tidak didukung.")
    
    try:
        dist_severitas = getattr(stats, nama_dist_severitas)
    except AttributeError:
        raise ValueError(f"Distribusi severitas {nama_dist_severitas} tidak ditemukan.")
    
    # Progress bar untuk simulasi
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for i in range(jumlah_iterasi):
        try:
            if nama_dist_frekuensi == 'nbinom':
                jumlah_klaim = int(dist_frekuensi.rvs(n=param_frekuensi['n'], p=param_frekuensi['p'], random_state=i))
            elif nama_dist_frekuensi == 'geom':
                jumlah_klaim = int(dist_frekuensi.rvs(p=param_frekuensi['p'], random_state=i))
            else:
                jumlah_klaim = int(dist_frekuensi.rvs(mu=param_frekuensi['mu'], random_state=i))
            
            if jumlah_klaim < 0:
                jumlah_klaim = 0
            
            klaim = dist_severitas.rvs(*param_severitas, size=jumlah_klaim, random_state=i) if jumlah_klaim > 0 else np.array([])
            
            for j in range(jumlah_klaim):
                severitas = max(int(klaim[j]), 0)
                data_tabel.append({
                    'Iterasi': f"{i+1}.{j+1}",
                    'Severitas': severitas,
                    'Flagging Frekuensi': jumlah_klaim if j == 0 else None
                })
            
            # Update progress
            if (i + 1) % max(1, jumlah_iterasi // 100) == 0:
                progress = (i + 1) / jumlah_iterasi
                progress_bar.progress(progress)
                status_text.text(f"Simulasi: {i+1}/{jumlah_iterasi} iterasi ({progress*100:.0f}%)")
                
        except Exception as e:
            st.warning(f"Iterasi {i+1} gagal: {str(e)}")
            continue
    
    progress_bar.empty()
    status_text.empty()
    
    if not data_tabel:
        raise ValueError("Simulasi gagal menghasilkan data.")
    
    df_tabel = pd.DataFrame(data_tabel)
    return df_tabel

@st.cache_data
def alokasikan_klaim(data_simulasi, ur, layer, data_iterasi=None, _cache_key=None):
    """Alokasi klaim ke UR dan layer"""
    hasil = []
    for idx, klaim in enumerate(data_simulasi):
        alokasi_klaim = {
            'Iterasi': data_iterasi.iloc[idx]['Iterasi'] if data_iterasi is not None else f"Real.{idx+1}",
            'Severitas': int(klaim),
            'UR': 0,
            'Layer 1': 0,
            'Layer 2': 0,
            'Layer 3': 0,
            'Layer 4': 0,
            'Layer 5': 0,
            'Layer 6': 0
        }
        sisa_klaim = max(0, klaim)
        
        alokasi_klaim['UR'] = min(sisa_klaim, ur)
        sisa_klaim -= alokasi_klaim['UR']
        
        for i, batas_layer in enumerate(layer, 1):
            if sisa_klaim <= 0:
                break
            alokasi_klaim[f'Layer {i}'] = min(sisa_klaim, batas_layer)
            sisa_klaim -= alokasi_klaim[f'Layer {i}']
        
        hasil.append(alokasi_klaim)
    
    return pd.DataFrame(hasil)

@st.cache_data
def rangkum_berdasarkan_frekuensi(df_simulasi, df_soc, jumlah_iterasi, _cache_key=None):
    """Rangkuman berdasarkan frekuensi"""
    data_frekuensi = df_simulasi.dropna(subset=['Flagging Frekuensi'])[['Iterasi', 'Flagging Frekuensi']]
    peta_frekuensi = data_frekuensi.set_index('Iterasi')['Flagging Frekuensi'].to_dict()
    
    ringkasan_severitas = df_simulasi.groupby(df_simulasi['Iterasi'].str.split('.').str[0]).agg({'Severitas': 'sum'}).rename(columns={'Severitas': 'Total Severitas'})
    ringkasan_severitas['Frekuensi'] = ringkasan_severitas.index.map(lambda x: peta_frekuensi.get(f"{int(x)}.1", 0))
    
    ringkasan_ur = df_soc.groupby(df_soc['Iterasi'].str.split('.').str[0]).agg({'UR': 'sum'}).rename(columns={'UR': 'Total UR'})
    ringkasan_ur['Frekuensi'] = ringkasan_ur.index.map(lambda x: peta_frekuensi.get(f"{int(x)}.1", 0))
    
    semua_iterasi = [str(i) for i in range(1, jumlah_iterasi + 1)]
    
    df_ringkasan = pd.DataFrame({
        'Frekuensi': [peta_frekuensi.get(f"{i}.1", 0) for i in semua_iterasi],
        'Total Severitas': [ringkasan_severitas.loc[str(i), 'Total Severitas'] if str(i) in ringkasan_severitas.index else 0 for i in semua_iterasi],
        'Total UR': [ringkasan_ur.loc[str(i), 'Total UR'] if str(i) in ringkasan_ur.index else 0 for i in semua_iterasi]
    }, index=semua_iterasi)
    
    df_ringkasan.index.name = 'Iterasi'
    return df_ringkasan.reset_index()

@st.cache_data
def rangkum_layer(df_soc, nomor_layer, batas_layer, jumlah_iterasi, maks_reinstatement, _cache_key=None):
    """Rangkuman layer tertentu"""
    ringkasan_layer = df_soc.groupby(df_soc['Iterasi'].str.split('.').str[0]).agg({
        f'Layer {nomor_layer}': 'sum',
    }).rename(columns={f'Layer {nomor_layer}': f'Total Layer {nomor_layer}'})
    
    frekuensi_layer = df_soc[df_soc[f'Layer {nomor_layer}'] > 0].groupby(
        df_soc['Iterasi'].str.split('.').str[0]
    ).size().to_frame(name=f'Frekuensi Layer {nomor_layer}')
    
    semua_iterasi = [str(i) for i in range(1, jumlah_iterasi + 1)]
    
    df_ringkasan = pd.DataFrame(index=semua_iterasi)
    df_ringkasan.index.name = 'Iterasi'
    df_ringkasan[f'Total Layer {nomor_layer}'] = [
        ringkasan_layer.loc[str(i), f'Total Layer {nomor_layer}'] if str(i) in ringkasan_layer.index else 0
        for i in semua_iterasi
    ]
    df_ringkasan[f'Frekuensi Layer {nomor_layer}'] = [
        frekuensi_layer.loc[str(i), f'Frekuensi Layer {nomor_layer}'] if str(i) in frekuensi_layer.index else 0
        for i in semua_iterasi
    ]
    
    for reinst in range(maks_reinstatement + 1):
        df_ringkasan[f'Reinstatement {reinst}'] = df_ringkasan.apply(
            lambda row: min(row[f'Total Layer {nomor_layer}'], (reinst + 1) * batas_layer),
            axis=1
        )
    
    return df_ringkasan.reset_index()

@st.cache_data
def hitung_premi(df_ringkasan_frekuensi, daftar_df_layer, layer, reinstatement_per_layer, _cache_key=None):
    """Hitung premi XoL"""
    daftar_df_premi = []
    maks_reinstatement = max(reinstatement_per_layer)
    
    for i, (df_layer, batas_layer, maks_reinst) in enumerate(zip(daftar_df_layer, layer, reinstatement_per_layer)):
        nomor_layer = i + 1
        if batas_layer <= 0:
            continue
        
        data_premi = {
            'Item': [f'Layer {nomor_layer}'],
            'Batas': [int(batas_layer)],
            'Rata-rata Klaim': [int(df_layer[f'Total Layer {nomor_layer}'].mean())],
            'Standar Deviasi Klaim': [int(df_layer[f'Total Layer {nomor_layer}'].std())],
            'Frekuensi Klaim': [int(df_layer[f'Frekuensi Layer {nomor_layer}'].sum())],
            'Total Klaim': [int(df_layer[f'Total Layer {nomor_layer}'].sum())]
        }
        df_premi = pd.DataFrame(data_premi)
        
        premi = []
        for reinst in range(maks_reinstatement + 1):
            if reinst > maks_reinst:
                premi.append(0)
                continue
            if f'Reinstatement {reinst}' not in df_layer.columns:
                premi.append(0)
                continue
            if reinst == 0:
                premi_reinst = df_layer[f'Reinstatement {reinst}'].mean()
            else:
                rata_rata_sekarang = df_layer[f'Reinstatement {reinst}'].mean()
                rata_rata_sebelumnya = df_layer[f'Reinstatement {reinst-1}'].mean()
                if rata_rata_sebelumnya > 0 and batas_layer > 0:
                    premi_reinst = rata_rata_sekarang / (1 + rata_rata_sebelumnya / batas_layer)
                else:
                    premi_reinst = 0
            premi.append(int(premi_reinst))
        
        for j, prem in enumerate(premi):
            df_premi[f'Reinstatement {j}'] = prem
        
        daftar_df_premi.append(df_premi)
    
    if not daftar_df_premi:
        raise ValueError("Tidak ada premi yang dapat dihitung.")
    
    df_premi_kombinasi = pd.concat(daftar_df_premi, ignore_index=True)
    
    total_premi = {
        'Item': 'Total',
        'Batas': '',
        'Rata-rata Klaim': int(df_premi_kombinasi['Rata-rata Klaim'].sum()),
        'Standar Deviasi Klaim': int(df_premi_kombinasi['Standar Deviasi Klaim'].sum()),
        'Frekuensi Klaim': int(df_premi_kombinasi['Frekuensi Klaim'].sum()),
        'Total Klaim': int(df_premi_kombinasi['Total Klaim'].sum())
    }
    for reinst in range(maks_reinstatement + 1):
        total_premi[f'Reinstatement {reinst}'] = int(df_premi_kombinasi[f'Reinstatement {reinst}'].sum())
    
    baris_total = pd.DataFrame([total_premi])
    df_premi_kombinasi = pd.concat([df_premi_kombinasi, baris_total], ignore_index=True)
    
    df_premi_kombinasi['Total'] = df_premi_kombinasi[[f'Reinstatement {i}' for i in range(maks_reinstatement + 1)]].sum(axis=1)
    
    return df_premi_kombinasi

def ringkasan_data_asli(df_soc_real, ur, layer):
    """Ringkasan data asli"""
    summary_data = []
    
    # Proses OR
    ur_data = df_soc_real['UR'].dropna()
    total_klaim_ur = int(ur_data.sum()) if not ur_data.empty else 0
    frekuensi_klaim_ur = int(len(ur_data[ur_data > 0]))
    rata_rata_klaim_ur = int(total_klaim_ur / frekuensi_klaim_ur) if frekuensi_klaim_ur > 0 else 0
    
    summary_data.append({
        'Item': 'OR',
        'Batas': int(ur),
        'Rata-rata Klaim (All Polis)': int(ur_data.mean()) if not ur_data.empty else 0,
        'Frekuensi Klaim': frekuensi_klaim_ur,
        'Total Klaim': total_klaim_ur,
        'Rata-rata Klaim per OR/Layer': rata_rata_klaim_ur
    })
    
    # Proses Layer 1 hingga 6
    for i in range(1, 7):
        layer_data = df_soc_real[f'Layer {i}'].dropna()
        total_klaim = int(layer_data.sum()) if not layer_data.empty else 0
        frekuensi_klaim = int(len(layer_data[layer_data > 0]))
        rata_rata_klaim = int(total_klaim / frekuensi_klaim) if frekuensi_klaim > 0 else 0
        
        summary_data.append({
            'Item': f'Layer {i}',
            'Batas': int(layer[i-1]),
            'Rata-rata Klaim (All Polis)': int(layer_data.mean()) if not layer_data.empty else 0,
            'Frekuensi Klaim': frekuensi_klaim,
            'Total Klaim': total_klaim,
            'Rata-rata Klaim per OR/Layer': rata_rata_klaim
        })
    
    df_summary = pd.DataFrame(summary_data)
    
    total_row = {
        'Item': 'Total',
        'Batas': '',
        'Rata-rata Klaim (All Polis)': int(df_summary['Rata-rata Klaim (All Polis)'].sum()),
        'Frekuensi Klaim': int(df_summary['Frekuensi Klaim'].sum()),
        'Total Klaim': int(df_summary['Total Klaim'].sum()),
        'Rata-rata Klaim per OR/Layer': int(df_summary['Total Klaim'].sum() / df_summary['Frekuensi Klaim'].sum()) if df_summary['Frekuensi Klaim'].sum() > 0 else 0
    }
    df_summary = pd.concat([df_summary, pd.DataFrame([total_row])], ignore_index=True)
    
    return df_summary

def create_distribution_chart(data, title="Distribusi Data"):
    """Buat chart distribusi data"""
    fig = go.Figure()
    fig.add_trace(go.Histogram(x=data, name='Histogram', nbinsx=50))
    fig.update_layout(
        title=title,
        xaxis_title="Nilai",
        yaxis_title="Frekuensi",
        showlegend=True,
        height=400
    )
    return fig

def create_excel_output(daftar_lembar):
    """Buat file Excel dengan styling"""
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    
    batas_tipis = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    perataan_tengah = Alignment(horizontal='center', vertical='center')
    format_rupiah = '#,##0'
    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for df, nama_lembar in daftar_lembar:
        ws = wb.create_sheet(title=nama_lembar)
        
        # Set tab color untuk sheet premi
        if nama_lembar == '1. Premi XoL':
            ws.sheet_properties.tabColor = "00205f"
        
        # Tulis data
        for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(r)
            
            # Style header
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = batas_tipis
                    cell.alignment = perataan_tengah
        
        # Style semua cell
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = batas_tipis
                cell.alignment = perataan_tengah
                if cell.column > 1 and isinstance(cell.value, (int, float)):
                    cell.number_format = format_rupiah
        
        # Auto-adjust column width
        for col in ws.columns:
            panjang_maks = 0
            kolom = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > panjang_maks:
                        panjang_maks = len(str(cell.value))
                except:
                    pass
            lebar_disesuaikan = min(panjang_maks + 4, 50)
            ws.column_dimensions[kolom].width = lebar_disesuaikan
        
        # Set row height
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 20
    
    wb.save(output)
    output.seek(0)
    return output

# ============================================================================
# APLIKASI UTAMA
# ============================================================================

def main():
    # Header
    st.markdown('<h1 class="main-header">📊 Pricing Excess of Loss dengan Reinstatement</h1>', unsafe_allow_html=True)
    
    show_info_box(" Aplikasi ini membantu menghitung premi Excess of Loss berdasarkan data historis klaim menggunakan simulasi Monte Carlo.")
    
    # Sidebar untuk navigasi dan informasi
    with st.sidebar:
        st.markdown("---")
        st.markdown("### 📋 Langkah-Langkah")
        st.markdown("""
        1. Upload data severitas & frekuensi
        2. Pilih kolom yang sesuai
        3. Lihat ringkasan data asli
        4. Atur parameter simulasi
        5. Jalankan simulasi
        6. Download hasil
        """)
    
    # Initialize session state
    if 'data_loaded' not in st.session_state:
        st.session_state.data_loaded = False
    if 'simulation_done' not in st.session_state:
        st.session_state.simulation_done = False
    
    # Tab Navigation
    tab1, tab2, tab3, tab4 = st.tabs(["📁 Upload Data", "📊 Analisis Data Asli", "⚙️ Simulasi", "📥 Hasil & Download"])
    
    # ========================================================================
    # TAB 1: UPLOAD DATA
    # ========================================================================
    with tab1:
        st.header("1️⃣ Upload Data Klaim")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📈 Data Severitas")
            file_severitas = st.file_uploader(
                "Upload file Excel berisi data severitas klaim",
                type=["xlsx", "xls"],
                key="severitas",
                help="File harus berformat Excel (.xlsx atau .xls)"
            )
            
        with col2:
            st.subheader("📊 Data Frekuensi")
            file_frekuensi = st.file_uploader(
                "Upload file Excel berisi data frekuensi klaim",
                type=["xlsx", "xls"],
                key="frekuensi",
                help="File harus berformat Excel (.xlsx atau .xls)"
            )
        
        if file_severitas and file_frekuensi:
            try:
                df_severitas = pd.read_excel(file_severitas)
                df_frekuensi = pd.read_excel(file_frekuensi)
                
                show_info_box(f" File berhasil dimuat! Severitas: {len(df_severitas)} baris, Frekuensi: {len(df_frekuensi)} baris", "success")
                
                st.subheader("Pilih Kolom Data")
                
                col1, col2 = st.columns(2)
                with col1:
                    kolom_severitas = st.selectbox(
                        "Pilih kolom untuk data severitas",
                        df_severitas.columns,
                        help="Pilih kolom yang berisi nilai klaim (severitas)"
                    )
                    st.dataframe(df_severitas[[kolom_severitas]].head(10), use_container_width=True)
                    
                with col2:
                    kolom_frekuensi = st.selectbox(
                        "Pilih kolom untuk data frekuensi",
                        df_frekuensi.columns,
                        help="Pilih kolom yang berisi jumlah klaim (frekuensi)"
                    )
                    st.dataframe(df_frekuensi[[kolom_frekuensi]].head(10), use_container_width=True)
                
                # Validasi dan proses data
                data_severitas = df_severitas[kolom_severitas].dropna().values
                data_frekuensi = df_frekuensi[kolom_frekuensi].dropna().values
                
                try:
                    data_severitas = data_severitas.astype(float)
                    data_frekuensi = data_frekuensi.astype(float)
                except ValueError:
                    st.error("❌ Kolom yang dipilih harus berisi data numerik.")
                    st.stop()
                
                # Validasi data
                errors = []
                if len(data_severitas) == 0:
                    errors.append("Data severitas kosong")
                if len(data_frekuensi) == 0:
                    errors.append("Data frekuensi kosong")
                if np.any(data_severitas <= 0):
                    errors.append("Data severitas harus berisi nilai positif")
                if np.any(data_frekuensi < 0):
                    errors.append("Data frekuensi tidak boleh negatif")
                if not np.all(data_frekuensi == data_frekuensi.astype(int)):
                    errors.append("Data frekuensi harus berupa bilangan bulat")
                
                if errors:
                    for error in errors:
                        st.error(f"❌ {error}")
                    st.stop()
                
                # Simpan data ke session state
                st.session_state.data_severitas = data_severitas
                st.session_state.data_frekuensi = data_frekuensi
                st.session_state.data_loaded = True
                
                show_info_box("✅ Data valid dan siap untuk dianalisis!", "success")
                
                # Statistik deskriptif
                st.subheader("📊 Statistik Deskriptif")
                col1, col2 = st.columns(2)
                
                with col1:
                    st.metric("📈 Severitas - Rata-rata", f"{int(data_severitas.mean()):,}")
                    st.metric("📈 Severitas - Median", f"{int(np.median(data_severitas)):,}")
                    st.metric("📈 Severitas - Std Dev", f"{int(data_severitas.std()):,}")
                    st.metric("📈 Severitas - Max", f"{int(data_severitas.max()):,}")
                    
                with col2:
                    st.metric("📊 Frekuensi - Rata-rata", f"{data_frekuensi.mean():.2f}")
                    st.metric("📊 Frekuensi - Median", f"{int(np.median(data_frekuensi))}")
                    st.metric("📊 Frekuensi - Std Dev", f"{data_frekuensi.std():.2f}")
                    st.metric("📊 Frekuensi - Max", f"{int(data_frekuensi.max())}")
                
                # Visualisasi distribusi
                st.subheader("📊 Visualisasi Distribusi Data")
                col1, col2 = st.columns(2)
                
                with col1:
                    fig_sev = create_distribution_chart(data_severitas, "Distribusi Severitas")
                    st.plotly_chart(fig_sev, use_container_width=True)
                    
                with col2:
                    fig_freq = create_distribution_chart(data_frekuensi, "Distribusi Frekuensi")
                    st.plotly_chart(fig_freq, use_container_width=True)
                
            except Exception as e:
                st.error(f"❌ Gagal membaca file Excel: {str(e)}")
                st.stop()
    
    # ========================================================================
    # TAB 2: ANALISIS DATA ASLI
    # ========================================================================
    with tab2:
        if not st.session_state.data_loaded:
            show_info_box("⚠️ Silakan upload data terlebih dahulu di tab 'Upload Data'", "warning")
        else:
            st.header("2️⃣ Analisis Data Klaim Asli")
            
            data_severitas = st.session_state.data_severitas
            data_frekuensi = st.session_state.data_frekuensi
            
            st.subheader("⚙️ Konfigurasi Layer")
            show_info_box("Tentukan batas Own Retention (OR) dan Layer untuk mengalokasikan klaim")
            
            with st.expander("🔧 Pengaturan OR dan Layer", expanded=True):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    ur = st.number_input(
                        "Own Retention (OR)",
                        min_value=0,
                        value=5000000000,
                        step=1000000000,
                        format="%d",
                        key="real_ur",
                        help="Batas maksimal klaim yang ditanggung sendiri"
                    )
                
                st.markdown("##### Layer Configuration")
                layer = []
                cols = st.columns(3)
                
                layer_defaults = [5000000000, 40000000000, 50000000000, 0, 0, 0]
                for i in range(6):
                    with cols[i % 3]:
                        batas = st.number_input(
                            f"Layer {i+1}",
                            min_value=0,
                            value=layer_defaults[i],
                            step=1000000,
                            format="%d",
                            key=f"real_layer_{i+1}",
                            help=f"Batas maksimal untuk Layer {i+1}"
                        )
                        layer.append(batas)
            
            # Spreading of Claim
            if st.button("🔄 Hitung Alokasi Klaim", type="primary", key="calc_soc_real"):
                with st.spinner("Menghitung spreading of claim..."):
                    df_soc_real = alokasikan_klaim(data_severitas, ur, layer)
                    st.session_state.df_soc_real = df_soc_real
                    st.session_state.ur_real = ur
                    st.session_state.layer_real = layer
                    show_info_box(" Alokasi klaim berhasil dihitung!", "success")
            
            if 'df_soc_real' in st.session_state:
                st.subheader("📋 Spreading of Claim (SoC) - Data Asli")
                st.dataframe(
                    st.session_state.df_soc_real.head(100),
                    hide_index=True,
                    use_container_width=True,
                    height=400
                )
                
                st.info(f"📊 Total data: {len(st.session_state.df_soc_real)} klaim")
                
                # Ringkasan
                st.subheader("📊 Ringkasan Alokasi Klaim")
                df_summary = ringkasan_data_asli(
                    st.session_state.df_soc_real,
                    st.session_state.ur_real,
                    st.session_state.layer_real
                )
                st.session_state.df_summary_real = df_summary
                
                st.dataframe(df_summary, hide_index=True, use_container_width=True)
                
                # Visualisasi
                st.subheader("📊 Visualisasi Alokasi")
                
                # Filter hanya layer yang aktif (batas > 0)
                active_layers = ['OR'] + [f'Layer {i+1}' for i in range(6) if st.session_state.layer_real[i] > 0]
                layer_totals = [df_summary[df_summary['Item'] == layer]['Total Klaim'].values[0] for layer in active_layers]
                
                fig = go.Figure(data=[
                    go.Bar(
                        x=active_layers,
                        y=layer_totals,
                        text=[f"{int(val):,}" for val in layer_totals],
                        textposition='auto',
                    )
                ])
                
                fig.update_layout(
                    title="Total Klaim per Layer",
                    xaxis_title="Layer",
                    yaxis_title="Total Klaim (Rp)",
                    showlegend=False,
                    height=500
                )
                
                st.plotly_chart(fig, use_container_width=True)
    
    # ========================================================================
    # TAB 3: SIMULASI
    # ========================================================================
    with tab3:
        if not st.session_state.data_loaded:
            show_info_box("Silakan upload data terlebih dahulu di tab 'Upload Data'", "warning")
        else:
            st.header("3️⃣ Simulasi Monte Carlo")
            
            data_severitas = st.session_state.data_severitas
            data_frekuensi = st.session_state.data_frekuensi
            
            # Parameter Frekuensi
            rata_rata_frekuensi = np.mean(data_frekuensi)
            varians_frekuensi = np.var(data_frekuensi)
            
            param_poisson = {'mu': rata_rata_frekuensi}
            param_negbinom = {
                'p': rata_rata_frekuensi / varians_frekuensi if varians_frekuensi > rata_rata_frekuensi else 0.99,
                'n': rata_rata_frekuensi ** 2 / (varians_frekuensi - rata_rata_frekuensi) if varians_frekuensi > rata_rata_frekuensi else rata_rata_frekuensi
            }
            param_geom = {'p': 1 / rata_rata_frekuensi if rata_rata_frekuensi > 0 else 0.99}
            
            st.subheader("📊 Parameter Distribusi Frekuensi")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                with st.container(border=True):
                    st.markdown("**Poisson**")
                    st.metric("μ (mu)", f"{param_poisson['mu']:.2f}")
            
            with col2:
                with st.container(border=True):
                    st.markdown("**Negative Binomial**")
                    st.metric("n", f"{param_negbinom['n']:.2f}")
                    st.metric("p", f"{param_negbinom['p']:.4f}")
            
            with col3:
                with st.container(border=True):
                    st.markdown("**Geometric**")
                    st.metric("p", f"{param_geom['p']:.4f}")
            
            # Fitting Distribusi Severitas
            st.subheader("📈 Fitting Distribusi Severitas")
            
            if 'metrik_severitas' not in st.session_state:
                with st.spinner("🔄 Menyesuaikan distribusi severitas... Mohon tunggu..."):
                    try:
                        cache_key_severity = str(uuid.uuid4())
                        metrik_severitas, param_severitas = sesuaikan_distribusi_severitas(
                            data_severitas,
                            _cache_key=cache_key_severity
                        )
                        st.session_state.metrik_severitas = metrik_severitas
                        st.session_state.param_severitas = param_severitas
                        show_info_box("Distribusi severitas berhasil disesuaikan!", "success")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        st.stop()
            
            metrik_severitas = st.session_state.metrik_severitas
            param_severitas = st.session_state.param_severitas
            
            st.markdown("**Top 10 Distribusi Terbaik (berdasarkan AIC):**")
            st.dataframe(metrik_severitas, hide_index=True, use_container_width=True)
            
            # Konfigurasi Simulasi
            st.subheader("⚙️ Konfigurasi Simulasi")
            
            col1, col2 = st.columns(2)
            
            with col1:
                jumlah_iterasi = st.number_input(
                    "Jumlah Iterasi",
                    min_value=100,
                    max_value=100000,
                    value=1000,
                    step=100,
                    help="Jumlah iterasi simulasi Monte Carlo (lebih banyak = lebih akurat tapi lebih lama)"
                )
                
                dist_frekuensi_pilih = st.selectbox(
                    "Pilih Distribusi Frekuensi",
                    ['poisson', 'nbinom', 'geom'],
                    format_func=lambda x: {
                        'poisson': 'Poisson',
                        'nbinom': 'Negative Binomial',
                        'geom': 'Geometric'
                    }[x]
                )
            
            with col2:
                dist_severitas_pilih = st.selectbox(
                    "Pilih Distribusi Severitas",
                    metrik_severitas['Distribusi'].values,
                    help="Pilih dari distribusi yang telah di-fit (diurutkan berdasarkan AIC terbaik)"
                )
                
                st.info(f"""
                **Distribusi Terpilih:**
                - Frekuensi: {dist_frekuensi_pilih}
                - Severitas: {dist_severitas_pilih}
                """)
            
            # Input Layer untuk Simulasi
            st.subheader("🎯 Konfigurasi OR dan Layer untuk Simulasi")
            
            with st.expander("🔧 Pengaturan Layer dan Reinstatement", expanded=True):
                ur_sim = st.number_input(
                    "Own Retention (OR)",
                    min_value=0,
                    value=5000000000,
                    step=1000000000,
                    format="%d",
                    key="sim_ur"
                )
                
                layer_sim = []
                reinstatement_per_layer = []
                
                st.markdown("##### Layer Configuration")
                
                for i in range(6):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        batas = st.number_input(
                            f"Layer {i+1}",
                            min_value=0,
                            value=5000000000 if i == 0 else 40000000000 if i == 1 else 50000000000 if i == 2 else 0,
                            step=1000000,
                            format="%d",
                            key=f"sim_layer_{i+1}"
                        )
                        layer_sim.append(batas)
                    
                    with col2:
                        reinst = st.number_input(
                            f"Reinstatement L{i+1}",
                            min_value=0,
                            max_value=100,
                            value=4 if batas > 0 else 0,
                            step=1,
                            key=f"sim_reinst_{i+1}",
                            help="Jumlah reinstatement untuk layer ini"
                        )
                        reinstatement_per_layer.append(reinst)
            
            # Tombol Jalankan Simulasi
            st.markdown("---")
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("🚀 Jalankan Simulasi Monte Carlo", type="primary", use_container_width=True):
                    try:
                        param_frekuensi = param_poisson if dist_frekuensi_pilih == 'poisson' else \
                                         param_negbinom if dist_frekuensi_pilih == 'nbinom' else param_geom
                        param_sev = param_severitas[dist_severitas_pilih]
                        
                        # Jalankan simulasi
                        cache_key_monte_carlo = str(uuid.uuid4())
                        df_tabel = jalankan_simulasi_monte_carlo(
                            jumlah_iterasi,
                            dist_frekuensi_pilih,
                            param_frekuensi,
                            dist_severitas_pilih,
                            param_sev,
                            _cache_key=cache_key_monte_carlo
                        )
                        
                        show_info_box(f"Simulasi selesai! Total klaim yang dihasilkan: {len(df_tabel)}", "success")
                        
                        # Alokasi klaim
                        cache_key_allocation = str(uuid.uuid4())
                        df_klaim = alokasikan_klaim(
                            df_tabel['Severitas'].values,
                            ur_sim,
                            layer_sim,
                            df_tabel,
                            _cache_key=cache_key_allocation
                        )
                        
                        # Ringkasan
                        cache_key_freq_summary = str(uuid.uuid4())
                        df_ringkasan_frekuensi = rangkum_berdasarkan_frekuensi(
                            df_tabel,
                            df_klaim,
                            jumlah_iterasi,
                            _cache_key=cache_key_freq_summary
                        )
                        
                        # Rangkum layer
                        daftar_df_layer = []
                        for i in range(6):
                            cache_key_layer = str(uuid.uuid4())
                            df_layer = rangkum_layer(
                                df_klaim,
                                i+1,
                                layer_sim[i],
                                jumlah_iterasi,
                                reinstatement_per_layer[i],
                                _cache_key=cache_key_layer
                            )
                            daftar_df_layer.append(df_layer)
                        
                        # Hitung premi
                        cache_key_premium = str(uuid.uuid4())
                        df_premi = hitung_premi(
                            df_ringkasan_frekuensi,
                            daftar_df_layer,
                            layer_sim,
                            reinstatement_per_layer,
                            _cache_key=cache_key_premium
                        )
                        
                        # Simpan ke session state
                        st.session_state.df_tabel = df_tabel
                        st.session_state.df_klaim = df_klaim
                        st.session_state.df_ringkasan_frekuensi = df_ringkasan_frekuensi
                        st.session_state.daftar_df_layer = daftar_df_layer
                        st.session_state.df_premi = df_premi
                        st.session_state.simulation_done = True
                        st.session_state.sim_params = {
                            'freq_dist': dist_frekuensi_pilih,
                            'sev_dist': dist_severitas_pilih,
                            'iterations': jumlah_iterasi,
                            'layer_sim': layer_sim,
                            'reinstatement': reinstatement_per_layer
                        }
                        
                        st.success("✅ Simulasi berhasil! Silakan lihat hasil di tab 'Hasil & Download'")
                        
                    except Exception as e:
                        st.error(f"❌ Error menjalankan simulasi: {str(e)}")
                        st.stop()
    
    # ========================================================================
    # TAB 4: HASIL & DOWNLOAD
    # ========================================================================
    with tab4:
        if not st.session_state.simulation_done:
            show_info_box("Silakan jalankan simulasi terlebih dahulu di tab 'Simulasi'", "warning")
        else:
            st.header("4️⃣ Hasil Simulasi dan Download")
            
            # Info simulasi
            params = st.session_state.sim_params
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Iterasi", f"{params['iterations']:,}")
            with col2:
                st.metric("Dist. Frekuensi", params['freq_dist'])
            with col3:
                st.metric("Dist. Severitas", params['sev_dist'])
            with col4:
                st.metric("Total Klaim", f"{len(st.session_state.df_tabel):,}")
            
            # Tabs untuk hasil
            result_tabs = st.tabs([
                "💰 Premi XoL",
                "📋 Hasil Simulasi",
                "📊 SoC",
                "📈 Klaim UR",
                "🔢 Layer Detail"
            ])
            
            # Tab Premi
            with result_tabs[0]:
                st.subheader("💰 Ringkasan Premi XoL")
                st.dataframe(
                    st.session_state.df_premi,
                    hide_index=True,
                    use_container_width=True
                )
                
                # Visualisasi premi per layer
                st.subheader("📊 Visualisasi Premi per Layer")
                
                df_viz = st.session_state.df_premi[st.session_state.df_premi['Item'] != 'Total'].copy()
                
                fig = go.Figure()
                
                for col in df_viz.columns:
                    if col.startswith('Reinstatement'):
                        fig.add_trace(go.Bar(
                            name=col,
                            x=df_viz['Item'],
                            y=df_viz[col],
                            text=[f"{int(val):,}" if val > 0 else "" for val in df_viz[col]],
                            textposition='auto'
                        ))
                
                fig.update_layout(
                    title="Premi per Layer dan Reinstatement",
                    xaxis_title="Layer",
                    yaxis_title="Premi (Rp)",
                    barmode='stack',
                    height=500,
                    showlegend=True
                )
                
                st.plotly_chart(fig, use_container_width=True)
            
            # Tab Hasil Simulasi
            with result_tabs[1]:
                st.subheader("📋 Hasil Simulasi Monte Carlo")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Klaim", f"{len(st.session_state.df_tabel):,}")
                with col2:
                    st.metric("Rata-rata Severitas", f"{int(st.session_state.df_tabel['Severitas'].mean()):,}")
                with col3:
                    st.metric("Median Severitas", f"{int(st.session_state.df_tabel['Severitas'].median()):,}")
                
                st.dataframe(
                    st.session_state.df_tabel.head(200),
                    hide_index=True,
                    use_container_width=True,
                    height=500
                )
            
            # Tab SoC
            with result_tabs[2]:
                st.subheader("📊 Spreading of Claim (SoC)")
                st.dataframe(
                    st.session_state.df_klaim.head(200),
                    hide_index=True,
                    use_container_width=True,
                    height=500
                )
            
            # Tab Klaim UR
            with result_tabs[3]:
                st.subheader("📈 Ringkasan Klaim UR")
                st.dataframe(
                    st.session_state.df_ringkasan_frekuensi.head(100),
                    hide_index=True,
                    use_container_width=True,
                    height=500
                )
            
            # Tab Layer Detail
            with result_tabs[4]:
                st.subheader("🔢 Detail per Layer")
                
                layer_tabs = st.tabs([f"Layer {i+1}" for i in range(6)])
                
                for i, layer_tab in enumerate(layer_tabs):
                    with layer_tab:
                        if params['layer_sim'][i] > 0:
                            df_layer_display = st.session_state.daftar_df_layer[i].drop(columns=["Iterasi"])
                            st.dataframe(
                                df_layer_display.head(100),
                                hide_index=True,
                                use_container_width=True,
                                height=400
                            )
                        else:
                            st.info(f"Layer {i+1} tidak aktif (batas = 0)")
            
            # Download Section
            st.markdown("---")
            st.subheader("📥 Download Hasil")
            
            show_info_box("File Excel akan berisi semua sheet hasil simulasi termasuk data asli dan hasil simulasi Monte Carlo")
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("📥 Generate & Download Excel", type="primary", use_container_width=True):
                    with st.spinner("🔄 Membuat file Excel..."):
                        try:
                            # Persiapkan data untuk Excel
                            daftar_lembar = [
                                (st.session_state.df_premi, '1. Premi XoL'),
                                (st.session_state.df_tabel, '2. Hasil Simulasi'),
                                (st.session_state.df_klaim, '3. Spreading of Claim'),
                                (st.session_state.df_ringkasan_frekuensi, '4. Klaim UR'),
                            ]
                            
                            # Tambahkan sheet untuk setiap layer
                            for i in range(6):
                                df_layer_display = st.session_state.daftar_df_layer[i].drop(columns=["Iterasi"])
                                daftar_lembar.append((df_layer_display, f'{5+i}. Layer {i+1}'))
                            
                            # Tambahkan data asli jika ada
                            if 'df_summary_real' in st.session_state:
                                daftar_lembar.insert(0, (st.session_state.df_summary_real, '0. Ringkasan Data Klaim'))
                            if 'df_soc_real' in st.session_state:
                                daftar_lembar.insert(1, (st.session_state.df_soc_real, '0.1. SoC (Data Klaim)'))
                            
                            # Buat Excel
                            output = create_excel_output(daftar_lembar)
                            
                            filename = f"Premi_XoL_Reinstatement_{params['freq_dist']}_{params['sev_dist']}.xlsx"
                            
                            st.download_button(
                                label="📥 Download File Excel",
                                data=output,
                                file_name=filename,
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                type="primary",
                                use_container_width=True
                            )
                            
                            show_info_box(f"✅ File {filename} siap diunduh!", "success")
                            
                        except Exception as e:
                            st.error(f"❌ Error membuat file Excel: {str(e)}")

if __name__ == "__main__":
    main()
